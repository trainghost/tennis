import os
import json 
import openpyxl
from flask import Flask, request, redirect, render_template, jsonify, url_for
from itertools import combinations
import sys


app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
DATA_FILE = "members_data.json"

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, encoding="utf-8") as f:
            return json.load(f)
    return []

def save_data(members):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(members, f, indent=2, ensure_ascii=False)

def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def best_match(members):
    min_diff = float("inf")
    best = ([], [])
    for comb in combinations(members, 4):
        for teamA in combinations(comb, 2):
            teamB = [m for m in comb if m not in teamA]
            diff = abs(sum(m["rank"] for m in teamA) - sum(m["rank"] for m in teamB))
            if diff < min_diff:
                min_diff = diff
                best = (list(teamA), list(teamB))
    return best

@app.route("/", methods=["GET", "POST"])
def index():
    members = load_data()
    for m in members:
        m.setdefault("late", False)

    if request.method == "POST":
        name = request.form["name"].strip()
        gender = request.form["gender"]
        rank = int(request.form["rank"] or 0)
        members.append({
            "name": name,
            "gender": gender,
            "rank": rank,
            "tuesday": False,
            "thursday": False,
            "participated": False
        })
        save_data(members)
        return redirect(url_for("index"))

    # GET 요청일 경우 정렬 적용
    sort = request.args.get("sort")

    if sort == "asc":
        members.sort(key=lambda x: x["rank"])
    elif sort == "desc":
        members.sort(key=lambda x: x["rank"], reverse=True)
    elif sort == "participated_first":
        members.sort(key=lambda x: not x.get("participated", False))  # 참여한 사람 먼저
    elif sort == "participated_last":
        members.sort(key=lambda x: x.get("participated", False))  # 참여 안 한 사람 먼저
    elif sort == "late_first":
        members.sort(key=lambda x: not x.get("late", False))
    elif sort == "late_last":
        members.sort(key=lambda x: x.get("late", False))


    return render_template("index.html", members=members, sort=sort)



@app.route("/toggle/<int:idx>/<key>")
def toggle(idx, key):
    members = load_data()
    if key in ["tuesday", "thursday", "participated", "late"]:
        members[idx][key] = not members[idx].get(key, False)
        save_data(members)
    return redirect(url_for("index"))
    
@app.route("/toggle/<int:idx>/<field>")
def toggle_field(idx, field):
    members = load_data()
    if 0 <= idx < len(members) and field in ["tuesday", "thursday", "participated", "late"]:  # ← 여기 "late" 포함되어야 함
        members[idx][field] = not members[idx].get(field, False)
        save_data(members)
    return redirect(url_for("index"))


@app.route("/delete/<int:idx>")
def delete(idx):
    members = load_data()
    if 0 <= idx < len(members):
        del members[idx]
    save_data(members)
    return redirect(url_for("index"))

@app.route("/upload", methods=["POST"])
def upload_excel():
    file = request.files["file"]
    if file and file.filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]
        members = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            members.append({
                "name": data.get("name", "").strip(),
                "gender": data.get("gender", "").strip(),
                "rank": int(data.get("rank", 0) or 0),
                "tuesday": bool(data.get("tuesday")),
                "thursday": bool(data.get("thursday")),
                "participated": bool(data.get("participated"))
            })

        save_data(members)
        return redirect(url_for("index"))

    return "올바른 엑셀 파일(.xlsx)을 업로드하세요.", 400


@app.route("/update_participation", methods=["POST"])
def update_participation():
    members = load_data()
    for i, member in enumerate(members):
        checkbox_name = f"participated_{i}"
        member["participated"] = checkbox_name in request.form
    save_data(members)
    return redirect(url_for("index"))

        
@app.route("/update_rank/<int:idx>", methods=["POST"])
def update_rank(idx):
    members = load_data()
    new_rank = int(request.form.get("rank", 0))
    if 0 <= idx < len(members):
        members[idx]["rank"] = new_rank
        save_data(members)
    return redirect(url_for("index"))

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filepath)

    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    new_members = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        gender = str(row[1]).strip() if len(row) > 1 and row[1] else "남"
        rank = int(row[2]) if len(row) > 2 and row[2] else 0

        new_members.append({
            "name": name,
            "gender": gender if gender in ["남", "여"] else "남",
            "rank": rank,
            "tuesday": False,
            "thursday": False,
            "participated": False
        })

    members = load_data()
    members.extend(new_members)
    save_data(members)

    return redirect(url_for("index"))


@app.route("/generate")
def generate_match():
    members = [m for m in load_data() if m.get("participated")]
    for m in members:
        m.setdefault("late", False)

    used_names = set()
    matches = []

    def make_match(title, candidate_pool, used_names):
        courts = []

        # 1. 사용된 사람 제외하고, 지각하지 않은 사람만
        remain = [m for m in candidate_pool if m["name"] not in used_names and not m.get("late")]

        # 2. 여자 코트 구성
        females = [m for m in remain if m["gender"] == "여"]
        if len(females) >= 4:
            A, B = best_match(females)
        elif len(females) >= 2:
            others = [m for m in remain if m["name"] not in [f["name"] for f in females]]
            mixed_candidates = females + others
            A, B = best_match([m for m in mixed_candidates if m["name"] not in used_names])
        else:
            candidates = [m for m in remain if m["name"] not in used_names]
            A, B = best_match(candidates)
        courts.append(("3번 코트", A, B))
        used_names.update(m["name"] for m in A + B)

        # 3. 남자 코트 구성
        males = [m for m in remain if m["gender"] == "남" and m["name"] not in used_names]
        males.sort(key=lambda x: x["rank"])

        if len(males) >= 8:
            A1 = [males[0], males[-1]]
            B1 = [males[1], males[-2]]
            courts.append(("4번 코트", A1, B1))
            used_names.update(m["name"] for m in A1 + B1)

            remain_males = [m for m in males if m["name"] not in used_names]
            if len(remain_males) >= 4:
                A2 = [remain_males[0], remain_males[-1]]
                B2 = [remain_males[1], remain_males[-2]]
                courts.append(("5번 코트", A2, B2))
                used_names.update(m["name"] for m in A2 + B2)

        return {
            "title": title,
            "courts": courts
        }

       # 매칭 1
    match1 = make_match("매칭 1 (06:10 ~ 06:35)", members, used_names)
    matches.append(match1)

        # 매칭 2 디버깅
    print("\n=== 매칭 1에 사용된 사람 ===")
    print(used_names)

    not_used = [m for m in members if m["name"] not in used_names]
    print("\n=== 매칭 2에서 우선 사용할 사람 (쉬는 사람) ===")
    print([m["name"] for m in not_used])

    fallback = [m for m in members if m["name"] in used_names]
    print("\n=== 매칭 2 보충용 후보 ===")
    print([m["name"] for m in fallback])

    match2_candidates = not_used + fallback

        # 매칭 2
    match2 = make_match("매칭 2 (06:37 ~ 07:00)", match2_candidates, used_names)
    matches.append(match2)


    return render_template("matches.html", matches=matches)





if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
